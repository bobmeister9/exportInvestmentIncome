import csv
import openpyxl
import argparse

class Transaction:
    Date = ''
    Acct = ''
    AcctType = ''
    Action = ''
    Security = ''
    Symbol = ''
    SecurityType = ''
    Category = ''
    Amt = 0
    Memo = ''

def readTransactions(filename):
    trans = []
    i = 0

    with open(filename, newline='') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if i > 0:
                # print(', '.join(row))
                t = Transaction()
                t.Date = row[0]
                t.Acct = row[1]
                t.Security = row[2]
                t.Action = row[3]
                t.Amt = float(row[5])
                if KeepTransaction(t):
                    trans.append(t)
            i += 1

    return trans

def KeepTransaction(t):
    # print (t.Acct, t.Date)
    if t.Amt < 0:
        return False
    if 'Revenue Credit' in t.Security:
        return True
    if t.Action == 'Retirement Contributions':
        return False
    if t.Action == 'Transfers':
        return False

    return True


def WriteHeader(sheet):
    sheet.cell(row=1, column=1).value = 'Date'
    sheet.cell(row=1, column=2).value = 'Account'
    sheet.cell(row=1, column=3).value = 'Account Type'
    sheet.cell(row=1, column=4).value = 'Action'
    sheet.cell(row=1, column=5).value = 'Security'
    sheet.cell(row=1, column=6).value = 'Symbol'
    sheet.cell(row=1, column=7).value = 'SecurityType'
    sheet.cell(row=1, column=8).value = 'Category'
    sheet.cell(row=1, column=9).value = 'Amount'

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input", help="Input file with transactions exported from Personal Capital", required=True)
    parser.add_argument("-o", "--output", help="output file name (.xlsx)", default="output.xlsx")

    args = parser.parse_args()

    transactions = readTransactions(args.input)

    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('Income')
    WriteHeader(sheet)
    nRow = 2 # skip the header
    for l in transactions:
        sheet.cell(row=nRow, column=1).value = l.Date
        sheet.cell(row=nRow, column=2).value = l.Acct
        sheet.cell(row=nRow, column=3).value = l.AcctType
        sheet.cell(row=nRow, column=4).value = l.Action
        sheet.cell(row=nRow, column=5).value = l.Security
        sheet.cell(row=nRow, column=6).value = l.Symbol
        sheet.cell(row=nRow, column=7).value = l.SecurityType
        sheet.cell(row=nRow, column=8).value = l.Category
        sheet.cell(row=nRow, column=9).value = l.Amt

        nRow += 1

    wb.save(args.output)
